"""Import libra nga Excel/CSV."""
import csv
import io
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import redirect
from openpyxl import load_workbook

from catalog.models import Author, Book, BookType, Copy, CopyStatus, Genre, Publisher, PurchaseMethod, Tag


def _norm(s):
    # Robust normalization: Excel/CSV cells can be numbers, booleans or None.
    return str(s or "").strip()


def _parse_authors(s):
    if not s:
        return []
    return [_norm(x) for x in str(s).replace(";", ",").split(",") if _norm(x)]


def _author_signature(names):
    return tuple(sorted({_norm(n).lower() for n in (names or []) if _norm(n)}))


def _row_signature(title, authors):
    return (_norm(title).lower(), _author_signature(authors))


def _get_or_create_authors(names):
    out = []
    for n in names:
        if not n:
            continue
        a, _ = Author.objects.get_or_create(name=n, defaults={"name": n})
        out.append(a)
    return out


def _get_or_create_genres(names):
    out = []
    for n in names:
        if not n:
            continue
        g = Genre.objects.filter(name__iexact=n).first()
        if not g:
            g = Genre.objects.create(name=n)
        out.append(g)
    return out


def _get_or_create_tags(names):
    out = []
    for n in names:
        if not n:
            continue
        t = Tag.objects.filter(name__iexact=n).first()
        if not t:
            t = Tag.objects.create(name=n)
        out.append(t)
    return out


def _get_or_create_publisher(name):
    if not name:
        return None
    p = Publisher.objects.filter(name__iexact=name).first()
    if not p:
        p = Publisher.objects.create(name=name)
    return p


def _get_val(row, keys, default=""):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip():
            return v
    return default


def _norm_header_key(k):
    return _norm(k).lower().replace(" ", "_")


def _has_required_title_header(headers):
    normalized = {_norm_header_key(h) for h in headers}
    return "titulli" in normalized


def _parse_purchase_method(raw):
    v = _norm(str(raw or "")).lower()
    if not v:
        return PurchaseMethod.FULL_PRICE
    if "donacion" in v:
        return PurchaseMethod.DONATION
    if "dhurat" in v:
        return PurchaseMethod.GIFT
    if "ulje" in v or "discount" in v:
        return PurchaseMethod.DISCOUNTED
    if "shkemb" in v or "këmbim" in v:
        return PurchaseMethod.EXCHANGE
    if "plot" in v or "full" in v or "cmim" in v or "çmim" in v:
        return PurchaseMethod.FULL_PRICE
    return PurchaseMethod.OTHER


def _extract_row_fields(row, is_excel=False):
    if is_excel:
        return {
            "titulli": _norm(_get_val(row, ["titulli", "Titulli"])),
            "isbn": _norm(_get_val(row, ["isbn", "ISBN"])),
            "pershkrimi": _norm(_get_val(row, ["pershkrimi", "përshkrimi", "Përshkrimi"])),
            "gjuha": _norm(_get_val(row, ["gjuha", "Gjuha"])),
            "viti": _get_val(row, ["viti", "Viti"]),
            "lloji": _norm(str(_get_val(row, ["lloji_librit", "lloji", "Lloji"]))),
            "botuesi": _norm(_get_val(row, ["botuesi", "Botuesi"])),
            "autoret": _parse_authors(_get_val(row, ["autoret", "Autorët", "autorët"])),
            "zhanret": _parse_authors(_get_val(row, ["zhanret", "Zhanret"])),
            "etiketa": _parse_authors(_get_val(row, ["etiketa", "Etiketa"])),
            "nr_kopjeve": _get_val(row, ["nr_kopjeve", "nr_kopjesh", "kopje"]),
            "lokacioni": _norm(_get_val(row, ["lokacioni", "location", "vendndodhja"])),
            "rafti": _norm(_get_val(row, ["rafti", "shelf", "rafte"])),
            "cmimi": _get_val(row, ["cmimi", "çmimi", "price", "cmim"]),
            "menyra_blerjes": _get_val(row, ["menyra_blerjes", "mënyra_blerjes", "menyra_e_blerjes"]),
            "vendi_blerjes": _norm(_get_val(row, ["vendi_blerjes", "vendi_i_blerjes", "vendi"])),
        }

    return {
        "titulli": _norm(row.get("titulli", "")),
        "isbn": _norm(row.get("isbn", "")),
        "pershkrimi": _norm(row.get("pershkrimi", "")),
        "gjuha": _norm(row.get("gjuha", "")),
        "viti": row.get("viti"),
        "lloji": _norm(row.get("lloji_librit", "")),
        "botuesi": _norm(row.get("botuesi", "")),
        "autoret": _parse_authors(row.get("autoret", "")),
        "zhanret": _parse_authors(row.get("zhanret", "")),
        "etiketa": _parse_authors(row.get("etiketa", "")),
        "nr_kopjeve": row.get("nr_kopjeve") or row.get("nr_kopjesh") or row.get("kopje"),
        "lokacioni": _norm(row.get("lokacioni", "") or row.get("location", "")),
        "rafti": _norm(row.get("rafti", "") or row.get("shelf", "")),
        "cmimi": row.get("cmimi") or row.get("çmimi") or row.get("price") or row.get("cmim"),
        "menyra_blerjes": row.get("menyra_blerjes") or row.get("mënyra_blerjes") or row.get("menyra_e_blerjes"),
        "vendi_blerjes": _norm(row.get("vendi_blerjes", "") or row.get("vendi_i_blerjes", "") or row.get("vendi", "")),
    }


def _is_duplicate_title_author_in_db(title, authors):
    sig = _row_signature(title, authors)
    candidates = Book.objects.filter(is_deleted=False, title__iexact=title).prefetch_related("authors")
    for b in candidates:
        db_sig = _row_signature(b.title, [a.name for a in b.authors.all()])
        if db_sig == sig:
            return True
    return False


def _import_row(row, is_excel=False):
    """Import one row. row is dict with keys."""
    fields = _extract_row_fields(row, is_excel=is_excel)
    titulli = fields["titulli"]
    isbn = fields["isbn"]
    pershkrimi = fields["pershkrimi"]
    gjuha = fields["gjuha"]
    viti = fields["viti"]
    lloji = fields["lloji"]
    botuesi = fields["botuesi"]
    autoret = fields["autoret"]
    zhanret = fields["zhanret"]
    etiketa = fields["etiketa"]
    nr_kopjeve = fields["nr_kopjeve"]
    lokacioni = fields["lokacioni"]
    rafti = fields["rafti"]
    cmimi = fields["cmimi"]
    menyra_blerjes = fields["menyra_blerjes"]
    vendi_blerjes = fields["vendi_blerjes"]

    if not titulli:
        return None, "Titulli zbrazët"

    try:
        viti_int = int(viti) if viti and str(viti).strip().isdigit() else None
    except (ValueError, TypeError):
        viti_int = None
    try:
        cmimi_val = str(cmimi or "").strip().replace(",", ".")
        price_decimal = Decimal(cmimi_val) if cmimi_val else None
    except (InvalidOperation, TypeError):
        price_decimal = None

    book_type = BookType.REFERENCE if lloji and "referenc" in lloji.lower() else BookType.GENERAL
    purchase_method = _parse_purchase_method(menyra_blerjes)
    publisher = _get_or_create_publisher(botuesi) if botuesi else None

    book = None
    created = False
    if isbn:
        book = Book.objects.filter(isbn=isbn, is_deleted=False).first()
    if book:
        book.title = titulli
        book.isbn = isbn or ""
        book.description = pershkrimi or ""
        book.language = gjuha or ""
        book.publication_year = viti_int
        book.book_type = book_type
        book.price = price_decimal
        book.purchase_method = purchase_method
        book.purchase_place = vendi_blerjes or ""
        book.publisher = publisher
        book.save()
        created = False
    else:
        book = Book.objects.create(
            title=titulli,
            isbn=isbn or "",
            description=pershkrimi or "",
            language=gjuha or "",
            publication_year=viti_int,
            book_type=book_type,
            price=price_decimal,
            purchase_method=purchase_method,
            purchase_place=vendi_blerjes or "",
            publisher=publisher,
        )
        created = True
    authors = _get_or_create_authors(autoret)
    genres = _get_or_create_genres(zhanret)
    tags = _get_or_create_tags(etiketa)
    book.authors.set(authors)
    book.genres.set(genres)
    book.tags.set(tags)

    # Krijo kopje nëse nr_kopjeve > 0
    try:
        n = int(nr_kopjeve) if nr_kopjeve and str(nr_kopjeve).strip().isdigit() else 0
    except (ValueError, TypeError):
        n = 0
    if n > 0:
        existing_count = book.copies.filter(is_deleted=False).count()
        for i in range(1, n + 1):
            barcode = f"IMP-{book.id}-{existing_count + i}"
            while Copy.objects.filter(barcode=barcode).exists():
                existing_count += 1
                barcode = f"IMP-{book.id}-{existing_count + i}"
            Copy.objects.create(
                book=book,
                barcode=barcode,
                status=CopyStatus.AVAILABLE,
                location=lokacioni or "",
                shelf=rafti or "",
            )

    return book, "Krijuar" if created else "Përditësuar"


@staff_member_required
def book_import(request):
    if request.method != "POST":
        return redirect("/admin/catalog/book/")

    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Zgjidhni një skedar Excel (.xlsx) ose CSV.")
        return redirect("/admin/catalog/book/")

    name = (f.name or "").lower()
    is_excel = name.endswith(".xlsx") or name.endswith(".xls")
    is_csv = name.endswith(".csv")

    if not (is_excel or is_csv):
        messages.error(request, "Formati duhet të jetë .xlsx ose .csv")
        return redirect("/admin/catalog/book/")

    created = 0
    updated = 0
    errors = []
    duplicate_rows = []
    parsed_rows = []

    try:
        if is_excel:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.warning(request, "Skedari është bosh.")
                return redirect("/admin/catalog/book/")
            raw_headers = [str(c or "") for c in rows[0]]
            if not _has_required_title_header(raw_headers):
                messages.error(
                    request,
                    "Skedari nuk ka kolonën e detyrueshme 'titulli'. Përdor skedarin shembull për formatin korrekt.",
                )
                return redirect("/admin/catalog/book/")

            headers = [_norm_header_key(c) for c in rows[0]]
            for idx, r in enumerate(rows[1:], start=2):
                row_dict = dict(zip(headers, r))
                # Skip fully empty rows
                if not any(_norm(v) for v in row_dict.values()):
                    continue
                parsed_rows.append((idx, row_dict, True))
        else:
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            if not reader.fieldnames or not _has_required_title_header(reader.fieldnames):
                messages.error(
                    request,
                    "CSV nuk ka kolonën e detyrueshme 'titulli'. Përdor skedarin shembull për formatin korrekt.",
                )
                return redirect("/admin/catalog/book/")
            for idx, row in enumerate(reader, start=2):
                row_clean = {_norm_header_key(k): v for k, v in (row or {}).items() if _norm(k)}
                if not any(_norm(v) for v in row_clean.values()):
                    continue
                parsed_rows.append((idx, row_clean, False))
    except Exception as e:
        messages.error(request, f"Gabim: {e}")
        return redirect("/admin/catalog/book/")

    if not parsed_rows:
        messages.warning(request, "Skedari nuk ka rreshta të vlefshëm për import.")
        return redirect("/admin/catalog/book/")

    # Validation pass: if there are errors, cancel entire import before writing.
    validation_errors = []
    valid_rows = []
    seen_signatures = set()
    for row_no, row_dict, row_is_excel in parsed_rows:
        fields = _extract_row_fields(row_dict, is_excel=row_is_excel)
        title = fields["titulli"]
        authors = fields["autoret"]
        if not title:
            validation_errors.append(f"Rreshti {row_no}: mungon vlera te kolona 'titulli'.")
            continue
        signature = _row_signature(title, authors)
        if signature in seen_signatures:
            duplicate_rows.append(f"Rreshti {row_no}: dublikatë brenda file-it ({title}).")
            continue
        if _is_duplicate_title_author_in_db(title, authors):
            duplicate_rows.append(f"Rreshti {row_no}: dublikatë ekzistuese në sistem ({title}).")
            continue
        seen_signatures.add(signature)
        valid_rows.append((row_no, row_dict, row_is_excel))

    if validation_errors:
        messages.error(
            request,
            f"Importi u anulua: u gjetën {len(validation_errors)} gabime në file. "
            "Korrigjo file-in dhe provo sërish.",
        )
        for err in validation_errors[:6]:
            messages.warning(request, err)
        return redirect("/admin/catalog/book/")

    # Import pass: process only after validation succeeded.
    for row_no, row_dict, row_is_excel in valid_rows:
        result, msg = _import_row(row_dict, is_excel=row_is_excel)
        if result:
            if msg == "Krijuar":
                created += 1
            else:
                updated += 1
        elif msg:
            errors.append(f"Rreshti {row_no}: {msg}")

    msg = f"Import u përfundua: {created} libra të rinj, {updated} përditësuar."
    if duplicate_rows:
        msg += f" {len(duplicate_rows)} rreshta u anashkaluan si dublikatë."
    if errors:
        msg += f" Gabime: {len(errors)}."
    messages.success(request, msg)
    if duplicate_rows:
        messages.warning(request, "Këta rreshta nuk u importuan sepse janë dublikatë (titull + autor):")
        for dup in duplicate_rows[:10]:
            messages.warning(request, dup)
    return redirect("/admin/catalog/book/")


@staff_member_required
def book_import_sample(request):
    """Shkarkon skedar shembull Excel me të gjitha fushat."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Libra"
    headers = [
        "titulli",
        "isbn",
        "pershkrimi",
        "gjuha",
        "viti",
        "lloji_librit",
        "botuesi",
        "autoret",
        "zhanret",
        "etiketa",
        "nr_kopjeve",
        "lokacioni",
        "rafti",
        "cmimi",
        "menyra_blerjes",
        "vendi_blerjes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    sample_rows = [
        [
            "Historia e Shqipërisë",
            "978-99943-1-234-5",
            "Një përmbledhje e plotë e historisë së Shqipërisë nga kohërat e lashta deri sot. Përfshin ngjarje kryesore, personalitete dhe zhvillime politike, ekonomike e kulturore.",
            "Shqip",
            2020,
            "I përgjithshëm",
            "Toena",
            "Kristo Frashëri, Skënder Anamali",
            "Histori, Edukativ",
            "shqiptare, referencë",
            3,
            "Raft A1",
            "A1-120",
            1200,
            "Blerje me çmim të plotë",
            "Toena Bookstore",
        ],
        [
            "Algoritmet dhe Strukturat e të Dhënave",
            "978-0-13-231681-1",
            "Tekst universitar për algoritme, struktura të dhënash dhe analizë kompleksiteti. Përfshin shembuj në Java dhe ushtrime.",
            "Anglisht",
            2012,
            "I përgjithshëm",
            "Pearson",
            "Robert Sedgewick",
            "Teknologji, Programim",
            "IT, CS",
            2,
            "Raft B2",
            "B2-045",
            850,
            "Blerje me ulje",
            "Panairi i Librit, Tiranë",
        ],
        [
            "Enciklopedia e Vogël",
            "",
            "Referencë e shpejtë për fakte të përgjithshme. Ideale për konsultim në bibliotekë.",
            "Shqip",
            2019,
            "Referencë (vetëm në bibliotekë)",
            "Botim i brendshëm",
            "Ekipi redaksional",
            "Referencë",
            "enciklopedi",
            1,
            "Sektori referencë",
            "REF-01",
            0,
            "Donacion",
            "Bashkia Kamëz",
        ],
        [
            "Rreth botës në 80 ditë",
            "978-99956-0-123-4",
            "Roman klasik i Jules Verne. Aventura e Phileas Fogg që vë bast që mund ta përfundojë rrethimin e globit në 80 ditë.",
            "Shqip",
            2018,
            "I përgjithshëm",
            "Dudaj",
            "Jules Verne",
            "Roman, Aventurë",
            "klasik, aventurë",
            5,
            "Raft C3",
            "C3-200",
            0,
            "Dhuratë",
            "Fondacion kulturor",
        ],
    ]
    for r, row in enumerate(sample_rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    col_widths = [28, 18, 45, 10, 8, 22, 18, 25, 18, 18, 12, 15, 12, 11, 20, 20]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    # Rresht udhëzimesh
    ws.cell(row=len(sample_rows) + 2, column=1, value="Udhëzime:")
    ws.cell(row=len(sample_rows) + 2, column=1).font = Font(bold=True, italic=True)
    tips = [
        "titulli: i detyrueshëm | isbn: opsional | pershkrimi: përshkrimi i librit",
        "autoret, zhanret, etiketa: ndani me presje (p.sh. Autor1, Autor2)",
        "nr_kopjeve: numri i kopjeve që do të krijohen (barkodet gjenerohen automatikisht)",
        "lokacioni, rafti: vendndodhja e kopjeve (opsionale)",
        "cmimi: numër (p.sh. 1200 ose 1200.50) | menyra_blerjes: Dhuratë, Donacion, Blerje me çmim të plotë, Blerje me ulje, Shkëmbim, Tjetër",
        "vendi_blerjes: dyqan/institucion/vend ku është siguruar libri",
    ]
    for i, tip in enumerate(tips, 1):
        ws.cell(row=len(sample_rows) + 2 + i, column=1, value=tip)
        ws.merge_cells(
            start_row=len(sample_rows) + 2 + i,
            start_column=1,
            end_row=len(sample_rows) + 2 + i,
            end_column=len(headers),
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="shembull_import_libra.xlsx"'
    wb.save(response)
    return response

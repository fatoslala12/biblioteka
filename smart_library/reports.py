"""Eksport raportesh në Excel dhe PDF — dizajn modern me ngjyra."""
import io
from datetime import date

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Paleta moderne
HEADER_FILL = "0d9488"      # teal-600
HEADER_FONT = "FFFFFF"
ROW_LIGHT = "f0fdfa"       # teal-50
ROW_DARK = "ccfbf1"        # teal-100
BORDER_COLOR = "94a3b8"     # slate-400
TITLE_COLOR = "0f766e"     # teal-700


def _excel_header_style():
    return {
        "font": Font(bold=True, color=HEADER_FONT, size=11),
        "fill": PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
            top=Side(style="thin", color=BORDER_COLOR),
            bottom=Side(style="thin", color=BORDER_COLOR),
        ),
    }


def _excel_cell_style(row_index):
    fill = ROW_LIGHT if row_index % 2 == 0 else ROW_DARK
    return {
        "fill": PatternFill(start_color=fill, end_color=fill, fill_type="solid"),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
            top=Side(style="thin", color=BORDER_COLOR),
            bottom=Side(style="thin", color=BORDER_COLOR),
        ),
    }


def _style_excel_sheet(ws, num_cols, num_rows):
    """Vë stilin e kokës dhe rreshtave në sheet."""
    header_style = _excel_header_style()
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        for k, v in header_style.items():
            setattr(cell, k, v)
    for r in range(2, num_rows + 2):
        row_style = _excel_cell_style(r)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            for k, v in row_style.items():
                setattr(cell, k, v)


def _member_display(m):
    if not m:
        return "", ""
    name = (m.full_name or "") or (m.user.get_full_name() if m.user_id else "") or m.member_no
    return name, m.member_no or ""


def export_loans_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Huazime"
    headers = [
        "ID", "Anëtari", "Nr. anëtarit", "Telefon anëtari", "Kopja", "Titulli", "ISBN", "Autorët", "Gjuha librit",
        "Marrë më", "Afati i kthimit", "Kthyer më", "Statusi", "Shënim",
        "Huazuar nga (staf)", "Dorëzuar nga (staf)", "Nr. zgjatjeve", "Krijuar më", "Përditësuar më",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    qs = queryset.select_related(
        "member", "member__user", "copy", "copy__book", "loaned_by", "returned_by"
    ).prefetch_related("copy__book__authors")
    for r, loan in enumerate(qs, 2):
        m = loan.member
        name, member_no = _member_display(m)
        book = loan.copy.book if loan.copy_id else None
        authors_str = ", ".join(a.name for a in (book.authors.all() if book else []))[:80] if book else ""
        ws.cell(row=r, column=1, value=loan.id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=member_no)
        ws.cell(row=r, column=4, value=m.phone if m else "")
        ws.cell(row=r, column=5, value=loan.copy.barcode if loan.copy_id else "")
        ws.cell(row=r, column=6, value=book.title if book else "")
        ws.cell(row=r, column=7, value=(book.isbn or "") if book else "")
        ws.cell(row=r, column=8, value=authors_str)
        ws.cell(row=r, column=9, value=(book.language or "") if book else "")
        ws.cell(row=r, column=10, value=loan.loaned_at.strftime("%Y-%m-%d %H:%M") if loan.loaned_at else "")
        ws.cell(row=r, column=11, value=loan.due_at.strftime("%Y-%m-%d") if loan.due_at else "")
        ws.cell(row=r, column=12, value=loan.returned_at.strftime("%Y-%m-%d %H:%M") if loan.returned_at else "")
        ws.cell(row=r, column=13, value=loan.get_status_display() if hasattr(loan, "get_status_display") else loan.status)
        ws.cell(row=r, column=14, value=(loan.note or "")[:100])
        ws.cell(row=r, column=15, value=loan.loaned_by.get_full_name() or loan.loaned_by.username if loan.loaned_by_id else "")
        ws.cell(row=r, column=16, value=loan.returned_by.get_full_name() or loan.returned_by.username if loan.returned_by_id else "")
        ws.cell(row=r, column=17, value=loan.renew_count or 0)
        ws.cell(row=r, column=18, value=loan.created_at.strftime("%Y-%m-%d %H:%M") if loan.created_at else "")
        ws.cell(row=r, column=19, value=loan.updated_at.strftime("%Y-%m-%d %H:%M") if loan.updated_at else "")
    _style_excel_sheet(ws, len(headers), queryset.count())
    col_widths = [5, 20, 10, 14, 12, 24, 16, 22, 10, 14, 12, 14, 10, 18, 14, 14, 8, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w, 50)
    return wb


def _pdf_table_style():
    return [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FONT}")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor(f"#{ROW_LIGHT}"), colors.HexColor(f"#{ROW_DARK}")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_COLOR}")),
        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor(f"#{TITLE_COLOR}")),
    ]


def _pdf_title(title_text):
    styles = getSampleStyleSheet()
    custom = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        textColor=colors.HexColor(f"#{TITLE_COLOR}"),
        fontSize=16,
        spaceAfter=6,
    )
    return Paragraph(title_text, custom)


def _pdf_subtitle():
    styles = getSampleStyleSheet()
    sub = ParagraphStyle(
        name="ReportSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.gray,
    )
    return Paragraph(f"Gjeneruar më {date.today().strftime('%d.%m.%Y')} — Smart Library", sub)


def export_loans_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=0.8 * cm, leftMargin=0.8 * cm,
        topMargin=1.2 * cm, bottomMargin=1 * cm,
    )
    headers = ["ID", "Anëtari", "Nr.", "Kopja", "Titulli", "Marrë", "Afati", "Kthyer", "Statusi", "Shënim"]
    data = [headers]
    qs = queryset.select_related("member", "member__user", "copy", "copy__book")[:300]
    for loan in qs:
        m = loan.member
        name, member_no = _member_display(m)
        data.append([
            str(loan.id),
            (name or "")[:16],
            (member_no or "")[:8],
            (loan.copy.barcode if loan.copy_id else "")[:10],
            (loan.copy.book.title if loan.copy_id else "")[:18],
            loan.loaned_at.strftime("%d.%m.%Y") if loan.loaned_at else "",
            loan.due_at.strftime("%d.%m.%Y") if loan.due_at else "",
            loan.returned_at.strftime("%d.%m.%Y") if loan.returned_at else "",
            str(loan.status)[:6],
            (loan.note or "")[:12],
        ])
    cw = [1.2 * cm, 3.2 * cm, 1.4 * cm, 2.4 * cm, 3.6 * cm, 2 * cm, 2 * cm, 2 * cm, 1.8 * cm, 2.2 * cm]
    t = Table(data, colWidths=cw)
    t.setStyle(TableStyle(_pdf_table_style()))
    doc.build([_pdf_title("Raport Huazimesh"), _pdf_subtitle(), Spacer(1, 0.3 * cm), t])
    return buffer.getvalue()


def export_reservations_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rezervime"
    headers = [
        "ID", "Anëtari", "Nr. anëtarit", "Telefon", "Libri", "ISBN", "Marrë më", "Kthim më", "Statusi",
        "Krijuar nga (staf)", "Huazuar nga (staf)", "Krijuar më", "Përditësuar më",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    qs = queryset.select_related("member", "member__user", "book", "created_by", "borrowed_by")
    for r, res in enumerate(qs, 2):
        m = res.member
        name, member_no = _member_display(m)
        ws.cell(row=r, column=1, value=res.id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=member_no)
        ws.cell(row=r, column=4, value=m.phone if m else "")
        ws.cell(row=r, column=5, value=res.book.title if res.book_id else "")
        ws.cell(row=r, column=6, value=(res.book.isbn or "") if res.book_id else "")
        ws.cell(row=r, column=7, value=res.pickup_date.strftime("%Y-%m-%d") if res.pickup_date else "")
        ws.cell(row=r, column=8, value=res.return_date.strftime("%Y-%m-%d") if res.return_date else "")
        ws.cell(row=r, column=9, value=res.get_status_display() if hasattr(res, "get_status_display") else res.status)
        ws.cell(row=r, column=10, value=res.created_by.get_full_name() or res.created_by.username if res.created_by_id else "")
        ws.cell(row=r, column=11, value=res.borrowed_by.get_full_name() or res.borrowed_by.username if res.borrowed_by_id else "")
        ws.cell(row=r, column=12, value=res.created_at.strftime("%Y-%m-%d %H:%M") if res.created_at else "")
        ws.cell(row=r, column=13, value=res.updated_at.strftime("%Y-%m-%d %H:%M") if res.updated_at else "")
    _style_excel_sheet(ws, len(headers), queryset.count())
    col_widths = [5, 20, 10, 14, 28, 16, 12, 12, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w, 50)
    return wb


def export_reservations_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=0.8 * cm, leftMargin=0.8 * cm,
        topMargin=1.2 * cm, bottomMargin=1 * cm,
    )
    headers = ["ID", "Anëtari", "Nr.", "Libri", "Marrë", "Kthim", "Statusi", "Krijuar më"]
    data = [headers]
    for res in queryset.select_related("member", "member__user", "book")[:300]:
        m = res.member
        name, member_no = _member_display(m)
        data.append([
            str(res.id),
            (name or "")[:16],
            (member_no or "")[:8],
            (res.book.title if res.book_id else "")[:22],
            res.pickup_date.strftime("%d.%m.%Y") if res.pickup_date else "",
            res.return_date.strftime("%d.%m.%Y") if res.return_date else "",
            str(res.status)[:8],
            res.created_at.strftime("%d.%m.%Y") if res.created_at else "",
        ])
    cw = [1.2 * cm, 3.5 * cm, 1.4 * cm, 4 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm]
    t = Table(data, colWidths=cw)
    t.setStyle(TableStyle(_pdf_table_style()))
    doc.build([_pdf_title("Raport Rezervimesh"), _pdf_subtitle(), Spacer(1, 0.3 * cm), t])
    return buffer.getvalue()


def export_fines_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gjoba"
    headers = [
        "ID", "Anëtari", "Nr. anëtarit", "ID huazimi", "Shuma", "Statusi", "Arsyeja",
        "Krijuar më", "Përditësuar më", "Falur nga (staf)", "Arsye falje",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    qs = queryset.select_related("member", "member__user", "loan", "waived_by")
    for r, fine in enumerate(qs, 2):
        m = fine.member
        name, member_no = _member_display(m)
        ws.cell(row=r, column=1, value=fine.id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=member_no)
        ws.cell(row=r, column=4, value=fine.loan_id if fine.loan_id else "")
        ws.cell(row=r, column=5, value=float(fine.amount))
        ws.cell(row=r, column=6, value=fine.get_status_display() if hasattr(fine, "get_status_display") else fine.status)
        ws.cell(row=r, column=7, value=(fine.reason or "")[:80])
        ws.cell(row=r, column=8, value=fine.created_at.strftime("%Y-%m-%d %H:%M") if fine.created_at else "")
        ws.cell(row=r, column=9, value=fine.updated_at.strftime("%Y-%m-%d %H:%M") if fine.updated_at else "")
        ws.cell(row=r, column=10, value=fine.waived_by.get_full_name() or fine.waived_by.username if fine.waived_by_id else "")
        ws.cell(row=r, column=11, value=(fine.waived_reason or "")[:50])
    _style_excel_sheet(ws, len(headers), queryset.count())
    col_widths = [5, 20, 10, 10, 10, 12, 28, 14, 14, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w, 50)
    return wb


def export_fines_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=0.8 * cm, leftMargin=0.8 * cm,
        topMargin=1.2 * cm, bottomMargin=1 * cm,
    )
    headers = ["ID", "Anëtari", "Nr.", "Shuma", "Statusi", "Arsyeja", "Falur nga", "Krijuar më"]
    data = [headers]
    for fine in queryset.select_related("member", "member__user", "waived_by")[:300]:
        m = fine.member
        name, member_no = _member_display(m)
        data.append([
            str(fine.id),
            (name or "")[:18],
            (member_no or "")[:8],
            str(fine.amount),
            str(fine.status)[:8],
            (fine.reason or "")[:16],
            (fine.waived_by.get_full_name() or fine.waived_by.username if fine.waived_by_id else "")[:12],
            fine.created_at.strftime("%d.%m.%Y") if fine.created_at else "",
        ])
    cw = [1.2 * cm, 3.5 * cm, 1.4 * cm, 2 * cm, 2 * cm, 3 * cm, 2.5 * cm, 2.2 * cm]
    t = Table(data, colWidths=cw)
    t.setStyle(TableStyle(_pdf_table_style()))
    doc.build([_pdf_title("Raport Gjobash"), _pdf_subtitle(), Spacer(1, 0.3 * cm), t])
    return buffer.getvalue()
